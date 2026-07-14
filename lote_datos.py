import openpyxl
from openpyxl.styles import Font, PatternFill

nombre_archivo = "Reporte_Masivo.xlsx"

lote_repuestos = [
    {"producto": "Bomba de Agua de Alta Presión", "cantidad": 45, "estatus": "OK"},
    {"producto": "Kit de Tiempo Completo Gate", "cantidad": 12, "estatus": "CRITICO"},
    {"producto": "Liga de Freno Wagner", "cantidad": 80, "estatus": "OK"}
]

libro = openpyxl.Workbook()
hoja = libro.active
hoja.title = "Lote Produccion"

fuente_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
fondo_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

fondo_alerta = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
fuente_alerta = Font(name="Arial", size=10, bold=True, color="78281F")

headers = ["Nombre del Repuesto", "Stock Actual", "Alerta de Sistema"]
hoja.append(headers)

for celda in hoja[1]:
    celda.font = fuente_header
    celda.fill = fondo_header

print("📊 Procesando lote con inteligencia visual condicional...")

for repuesto in lote_repuestos:
    estatus_actual = repuesto.get("estatus")
    
    hoja.append([repuesto.get("producto"), repuesto.get("cantidad"), estatus_actual])
    
    if estatus_actual == "CRITICO":
        fila_actual = hoja.max_row
        celda_alerta = hoja.cell(row=fila_actual, column=3)
        
        celda_alerta.fill = fondo_alerta
        celda_alerta.font = fuente_alerta

for columna in hoja.columns:
    letra_columna = columna[0].column_letter
    largo_maximo = max(len(str(celda.value or '')) for celda in columna)
    hoja.column_dimensions[letra_columna].width = largo_maximo + 4

libro.save(nombre_archivo)
print(f"🎨 ¡SISTEMA COMPLETADO! Alertas visuales inyectadas con éxito en '{nombre_archivo}'.")