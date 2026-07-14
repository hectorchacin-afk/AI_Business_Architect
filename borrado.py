import openpyxl

producto_a_eliminar = "Liga de Freno Wagner"
nombre_archivo = "Reporte_Masivo.xlsx"

try:
    print(f"🧹 Iniciando protocolo de depuración para: '{producto_a_eliminar}'...")
    libro = openpyxl.load_workbook(nombre_archivo)
    hoja = libro["Lote Produccion"]
    
    encontrado = False
    
    for num_fila in range(2, hoja.max_row + 1):
        nombre_celda = hoja.cell(row=num_fila, column=1).value
        
        if nombre_celda == producto_a_eliminar:
            print(f"🎯 Registro localizado en la fila número {num_fila}.")
            
            hoja.delete_rows(idx=num_fila, amount=1)
            
            print(f"🗑️ Fila {num_fila} eliminada físicamente de la tabla.")
            encontrado = True
            break
            
    if encontrado:
        libro.save(nombre_archivo)
        print(f"✅ ¡OPERACIÓN DEPURADA! El archivo '{nombre_archivo}' ha sido actualizado.")
    else:
        print(f"❌ ALERTA: El producto '{producto_a_eliminar}' no se encontró en la tabla.")

except FileNotFoundError:
    print(f"🚨 ERROR: No se encontró el archivo '{nombre_archivo}' para ejecutar la depuración.")