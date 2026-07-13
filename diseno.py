import openpyxl
from openpyxl.styles import Font, PatternFill

nombre_archivo = "Reporte_Diseno.xlsx"

libro = openpyxl.Workbook()
hoja = libro.active
hoja.title = "Estilo Ejecutivo"

hoja["A1"] = "Modulo Automatizado"
hoja["B1"] = "Estatus de Red"

fuente_premium = Font(name="Arial", size=12, bold=True, color="FFFFFF")

fondo_corporativo = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

hoja["A1"].font = fuente_premium
hoja["A1"].fill = fondo_corporativo

hoja["B1"].font = fuente_premium
hoja["B1"].fill = fondo_corporativo

hoja["A2"] = "Bot de Ciberseguridad"
hoja["B2"] = "ACTIVO"

libro.save(nombre_archivo)

print("🎨 ¡HITO DE DISEÑO! El archivo 'Reporte_Diseno.xlsx' ha sido creado con estilos corporativos.")