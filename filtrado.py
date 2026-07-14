import openpyxl
from openpyxl.styles import Font, PatternFill

try:
    print("🔍 Cargando base de datos maestra para filtrado premium...")
    libro_maestro = openpyxl.load_workbook("Reporte_Masivo.xlsx")
    hoja_maestra = libro_maestro["Lote Produccion"]
    
    libro_alerta = openpyxl.Workbook()
    hoja_alerta = libro_alerta.active
    hoja_alerta.title = "PRODUCTOS EN CRITICO"
    
    fondo_rojo = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    fuente_roja = Font(name="Arial", size=10, bold=True, color="78281F")
    
    hoja_alerta.append(["Repuesto de Emergencia", "Stock Actual"])
    
    print("⚡ Analizando y estilizando celdas en tiempo real...")
    
    for fila in hoja_maestra.iter_rows(min_row=2, max_row=4, values_only=True):
        nombre = fila
        stock = fila
        estatus = fila
        
        if estatus == "CRITICO":

            hoja_alerta.append([nombre, stock])
            
            fila_nueva = hoja_alerta.max_row
            
            celda_stock = hoja_alerta.cell(row=fila_nueva, column=2)
            celda_stock.fill = fondo_rojo
            celda_stock.font = fuente_roja
            
            print(f"🚨 Alerta aislada y estilizada: {nombre} ({stock} unidades)")
            
    libro_alerta.save("Lista_Compras_Urgentes.xlsx")
    print("\n💾 ¡SIMULACRO EXITOSO! Reporte estilizado generado con éxito.")

except FileNotFoundError:
    print("🚨 ERROR: No se encontró el archivo base para realizar el filtro.")